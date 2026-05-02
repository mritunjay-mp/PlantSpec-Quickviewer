import math
import os
import re
from collections import deque
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from PIL import Image, ImageDraw, ImageFont


WORKSPACE = Path.cwd()
BASE = Path(os.environ.get("MICASENSE_DIR", r"C:\Users\totso\OneDrive\MFM\R&D\실험\인도현지실험\micasense"))
OUT = WORKSPACE / "paper_figure_outputs" / "labeled_multispectral"
PREVIEW = OUT / "previews"
FIGURES = OUT / "figures"
STEPS = OUT / "step_figures"
SEGMENTS = OUT / "plant_segments"

BANDS = {
    1: "blue",
    2: "green",
    3: "red",
    4: "nir1",
    5: "nir2_red_edge",
    7: "thermal",
}


def ensure_dirs():
    OUT.mkdir(parents=True, exist_ok=True)
    PREVIEW.mkdir(parents=True, exist_ok=True)
    FIGURES.mkdir(parents=True, exist_ok=True)
    STEPS.mkdir(parents=True, exist_ok=True)
    SEGMENTS.mkdir(parents=True, exist_ok=True)


def read_tif(path, target_shape=None):
    arr = np.asarray(Image.open(path), dtype=np.float32)
    if target_shape and arr.shape != target_shape:
        img = Image.fromarray(arr)
        img = img.resize((target_shape[1], target_shape[0]), Image.Resampling.BILINEAR)
        arr = np.asarray(img, dtype=np.float32).copy()
    arr[arr <= 0] = np.nan
    return arr


def normalize_for_alignment(arr):
    a = arr.copy()
    finite = np.isfinite(a)
    if finite.sum() == 0:
        return np.zeros_like(a, dtype=np.float32)
    lo, hi = np.nanpercentile(a[finite], [2, 98])
    if hi <= lo:
        return np.zeros_like(a, dtype=np.float32)
    a = np.clip((a - lo) / (hi - lo), 0, 1)
    a[~np.isfinite(a)] = 0
    return a.astype(np.float32)


def estimate_shift(reference, moving, max_shift=80, sample=4):
    ref = normalize_for_alignment(reference)[::sample, ::sample]
    mov = normalize_for_alignment(moving)[::sample, ::sample]
    ref = ref - ref.mean()
    mov = mov - mov.mean()
    corr = np.fft.ifft2(np.fft.fft2(ref) * np.conj(np.fft.fft2(mov)))
    y, x = np.unravel_index(np.argmax(np.abs(corr)), corr.shape)
    if y > ref.shape[0] // 2:
        y -= ref.shape[0]
    if x > ref.shape[1] // 2:
        x -= ref.shape[1]
    shift_y, shift_x = int(y * sample), int(x * sample)
    shift_y = int(np.clip(shift_y, -max_shift, max_shift))
    shift_x = int(np.clip(shift_x, -max_shift, max_shift))
    return shift_y, shift_x


def shift_with_nan(arr, shift_y, shift_x):
    out = np.full_like(arr, np.nan)
    src_y0 = max(0, -shift_y)
    src_y1 = min(arr.shape[0], arr.shape[0] - shift_y)
    dst_y0 = max(0, shift_y)
    dst_y1 = min(arr.shape[0], arr.shape[0] + shift_y)
    src_x0 = max(0, -shift_x)
    src_x1 = min(arr.shape[1], arr.shape[1] - shift_x)
    dst_x0 = max(0, shift_x)
    dst_x1 = min(arr.shape[1], arr.shape[1] + shift_x)
    if src_y1 > src_y0 and src_x1 > src_x0:
        out[dst_y0:dst_y1, dst_x0:dst_x1] = arr[src_y0:src_y1, src_x0:src_x1]
    return out


def center_crop(arr, crop_upper_bound=None):
    if crop_upper_bound is None:
        crop_upper_bound = float(os.environ.get("CROP_MARGIN", "0.80"))
    fraction_each_side = 1.0 - crop_upper_bound
    h, w = arr.shape
    y0, y1 = int(h * fraction_each_side), int(h * (1 - fraction_each_side))
    x0, x1 = int(w * fraction_each_side), int(w * (1 - fraction_each_side))
    return arr[y0:y1, x0:x1]


def analysis_sample(arr):
    stride = int(os.environ.get("PROCESS_STRIDE", "4"))
    if stride <= 1:
        return arr
    return arr[::stride, ::stride]


def otsu_threshold(values, bins=128):
    values = values[np.isfinite(values)]
    if values.size == 0:
        return np.nan
    lo, hi = np.nanpercentile(values, [1, 99])
    if hi <= lo:
        return float(np.nanmean(values))
    hist, edges = np.histogram(np.clip(values, lo, hi), bins=bins, range=(lo, hi))
    centers = (edges[:-1] + edges[1:]) / 2
    weight1 = np.cumsum(hist)
    weight2 = np.cumsum(hist[::-1])[::-1]
    mean1 = np.cumsum(hist * centers) / np.maximum(weight1, 1)
    mean2 = (np.cumsum((hist * centers)[::-1]) / np.maximum(weight2[::-1], 1))[::-1]
    variance = weight1[:-1] * weight2[1:] * (mean1[:-1] - mean2[1:]) ** 2
    if variance.size == 0:
        return float(np.nanmean(values))
    return float(centers[:-1][np.argmax(variance)])


def green_mask(blue, green, red):
    b = normalize_for_alignment(blue)
    g = normalize_for_alignment(green)
    r = normalize_for_alignment(red)
    exg = 2 * g - r - b
    threshold = otsu_threshold(exg)
    mask = (
        np.isfinite(blue)
        & np.isfinite(green)
        & np.isfinite(red)
        & (exg > threshold)
        & (g > np.nanpercentile(g[np.isfinite(g)], 35))
        & (g >= r * 0.92)
        & (g >= b * 0.92)
    )
    return remove_small(mask, min_area=max(80, int(mask.size * 0.001)))


def remove_small(mask, min_area):
    labels, comps = connected_components(mask, min_area=min_area)
    return labels > 0


def connected_components(mask, min_area=100):
    h, w = mask.shape
    labels = np.zeros((h, w), dtype=np.int32)
    comps = []
    current = 0
    for yy in range(h):
        for xx in range(w):
            if not mask[yy, xx] or labels[yy, xx]:
                continue
            q = deque([(yy, xx)])
            labels[yy, xx] = -1
            pixels = []
            while q:
                y, x = q.popleft()
                pixels.append((y, x))
                for dy, dx in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                    ny, nx = y + dy, x + dx
                    if 0 <= ny < h and 0 <= nx < w and mask[ny, nx] and labels[ny, nx] == 0:
                        labels[ny, nx] = -1
                        q.append((ny, nx))
            if len(pixels) >= min_area:
                current += 1
                ys = [p[0] for p in pixels]
                xs = [p[1] for p in pixels]
                for y, x in pixels:
                    labels[y, x] = current
                comps.append(
                    {
                        "region_id": f"object_{current:02d}",
                        "area_px": len(pixels),
                        "bbox_y0": min(ys),
                        "bbox_x0": min(xs),
                        "bbox_y1": max(ys),
                        "bbox_x1": max(xs),
                    }
                )
            else:
                for y, x in pixels:
                    labels[y, x] = 0
    return labels, comps


def grid_regions(mask, n=3):
    h, w = mask.shape
    labels = np.zeros((h, w), dtype=np.int32)
    comps = []
    rid = 0
    min_area = max(50, int(mask.size * 0.0005))
    for gy in range(n):
        for gx in range(n):
            y0, y1 = int(h * gy / n), int(h * (gy + 1) / n)
            x0, x1 = int(w * gx / n), int(w * (gx + 1) / n)
            cell = mask[y0:y1, x0:x1]
            area = int(cell.sum())
            if area < min_area:
                continue
            rid += 1
            sub = labels[y0:y1, x0:x1]
            sub[cell] = rid
            comps.append(
                {
                    "region_id": f"grid_{gy + 1}_{gx + 1}",
                    "area_px": area,
                    "bbox_y0": y0,
                    "bbox_x0": x0,
                    "bbox_y1": y1 - 1,
                    "bbox_x1": x1 - 1,
                }
            )
    return labels, comps


def choose_regions(mask):
    labels = np.zeros(mask.shape, dtype=np.int32)
    labels[mask] = 1
    ys, xs = np.where(mask)
    if ys.size == 0:
        comps = []
    else:
        comps = [
            {
                "region_id": "plant_area",
                "area_px": int(mask.sum()),
                "bbox_y0": int(ys.min()),
                "bbox_x0": int(xs.min()),
                "bbox_y1": int(ys.max()),
                "bbox_x1": int(xs.max()),
            }
        ]
    return "plant_mask", labels, comps


def safe_index(num, den):
    out = num / np.where(np.abs(den) < 1e-6, np.nan, den)
    return np.clip(out, -2, 2)


def thermal_to_celsius(raw):
    # MicaSense LWIR is commonly stored as Kelvin x 100 in these TIFFs.
    return raw / 100.0 - 273.15


def stats(arr, mask):
    vals = arr[mask & np.isfinite(arr)]
    if vals.size == 0:
        return {"n_px": 0, "mean": np.nan, "sd": np.nan, "median": np.nan, "p10": np.nan, "p90": np.nan}
    return {
        "n_px": int(vals.size),
        "mean": float(np.nanmean(vals)),
        "sd": float(np.nanstd(vals)),
        "median": float(np.nanmedian(vals)),
        "p10": float(np.nanpercentile(vals, 10)),
        "p90": float(np.nanpercentile(vals, 90)),
    }


def stretch(arr):
    vals = arr[np.isfinite(arr)]
    if vals.size == 0:
        return np.zeros(arr.shape, dtype=np.uint8)
    lo, hi = np.nanpercentile(vals, [2, 98])
    if hi <= lo:
        hi = lo + 1
    return np.uint8(np.clip((arr - lo) / (hi - lo), 0, 1) * 255)


def heatmap(arr, vmin=-0.2, vmax=0.9):
    s = np.clip((arr - vmin) / (vmax - vmin), 0, 1)
    s[~np.isfinite(arr)] = 0
    r = np.uint8(np.clip(1.5 * s - 0.2, 0, 1) * 255)
    g = np.uint8(np.clip(1.8 * (1 - np.abs(s - 0.55) / 0.55), 0, 1) * 255)
    b = np.uint8(np.clip(1.2 * (1 - s), 0, 1) * 255)
    return Image.fromarray(np.dstack([r, g, b]), "RGB")


def save_preview(crop, label, capture_id, bands, mask, labels, ndvi, cwsi):
    rgb = np.dstack([stretch(bands["red"]), stretch(bands["green"]), stretch(bands["blue"])])
    rgb_img = Image.fromarray(rgb, "RGB")
    overlay = rgb_img.copy()
    draw = ImageDraw.Draw(overlay)
    mask_img = np.asarray(overlay).copy()
    mask_img[mask] = (0.45 * mask_img[mask] + 0.55 * np.array([40, 220, 80])).astype(np.uint8)
    overlay = Image.fromarray(mask_img, "RGB")
    draw = ImageDraw.Draw(overlay)
    for rid in np.unique(labels):
        if rid <= 0:
            continue
        ys, xs = np.where(labels == rid)
        if ys.size:
            draw.rectangle([int(xs.min()), int(ys.min()), int(xs.max()), int(ys.max())], outline=(255, 255, 0), width=2)
            draw.text((int(xs.min()) + 3, int(ys.min()) + 3), str(rid), fill=(255, 255, 0), font=ImageFont.load_default())

    panels = [
        ("RGB + green mask/regions", overlay),
        ("NDVI", heatmap(ndvi)),
        ("CWSI absolute", heatmap(cwsi, 0, 1)),
    ]
    resized = []
    for title, img in panels:
        im = img.resize((420, 315))
        d = ImageDraw.Draw(im)
        d.rectangle([0, 0, 420, 24], fill=(255, 255, 255))
        d.text((6, 7), title, fill=(0, 0, 0), font=ImageFont.load_default())
        resized.append(im)
    canvas = Image.new("RGB", (1260, 315), "white")
    for i, im in enumerate(resized):
        canvas.paste(im, (420 * i, 0))
    out = PREVIEW / f"{crop}_{label}_{capture_id}_preview.png"
    canvas.save(out)
    return out


def text_panel(title, lines, size=(420, 315)):
    img = Image.new("RGB", size, "white")
    draw = ImageDraw.Draw(img)
    font = ImageFont.load_default()
    draw.rectangle([0, 0, size[0], 28], fill=(235, 242, 235))
    draw.text((8, 9), title, fill=(0, 0, 0), font=font)
    y = 48
    for line in lines:
        draw.text((10, y), str(line), fill=(0, 0, 0), font=font)
        y += 18
    return img


def compose_panels(panels, out_path, panel_size=(420, 315), columns=3):
    rows = math.ceil(len(panels) / columns)
    canvas = Image.new("RGB", (panel_size[0] * columns, panel_size[1] * rows), "white")
    for i, (title, img) in enumerate(panels):
        im = img.resize(panel_size)
        draw = ImageDraw.Draw(im)
        draw.rectangle([0, 0, panel_size[0], 24], fill=(255, 255, 255))
        draw.text((6, 7), title, fill=(0, 0, 0), font=ImageFont.load_default())
        canvas.paste(im, ((i % columns) * panel_size[0], (i // columns) * panel_size[1]))
    canvas.save(out_path)
    return out_path


def mask_overlay(rgb_img, mask):
    base = np.asarray(rgb_img).copy()
    base[mask] = (0.45 * base[mask] + 0.55 * np.array([30, 220, 70])).astype(np.uint8)
    return Image.fromarray(base, "RGB")


def save_step_figures(crop, label, capture_id, raw_bands, aligned_full, crop_bands, mask, indices, cwsi, shifts):
    prefix = f"{crop}_{label}_{capture_id}"
    raw_rgb = Image.fromarray(
        np.dstack([stretch(raw_bands["red"]), stretch(raw_bands["green"]), stretch(raw_bands["blue"])]),
        "RGB",
    )
    aligned_rgb = Image.fromarray(
        np.dstack([stretch(aligned_full["red"]), stretch(aligned_full["green"]), stretch(aligned_full["blue"])]),
        "RGB",
    )
    crop_rgb = Image.fromarray(
        np.dstack([stretch(crop_bands["red"]), stretch(crop_bands["green"]), stretch(crop_bands["blue"])]),
        "RGB",
    )
    exg = 2 * normalize_for_alignment(crop_bands["green"]) - normalize_for_alignment(crop_bands["red"]) - normalize_for_alignment(crop_bands["blue"])
    thermal_c = indices["Thermal_C"]

    alignment = compose_panels(
        [
            ("01 raw RGB", raw_rgb),
            ("02 aligned RGB", aligned_rgb),
            (
                "alignment shifts",
                text_panel(
                    "Green reference shifts",
                    [
                        f"blue: {shifts.get('blue')}",
                        f"red: {shifts.get('red')}",
                        f"nir1: {shifts.get('nir1')}",
                        f"nir2/red-edge: {shifts.get('nir2_red_edge')}",
                        f"thermal: {shifts.get('thermal')}",
                    ],
                ),
            ),
        ],
        STEPS / f"{prefix}_01_alignment.png",
    )
    segmentation = compose_panels(
        [
            ("03 center crop RGB", crop_rgb),
            ("04 Green/ExG response", heatmap(exg, np.nanpercentile(exg, 2), np.nanpercentile(exg, 98))),
            ("05 plant mask overlay", mask_overlay(crop_rgb, mask)),
        ],
        STEPS / f"{prefix}_02_crop_mask.png",
    )
    indices_fig = compose_panels(
        [
            ("06 NDVI", heatmap(indices["NDVI"])),
            ("07 GNDVI", heatmap(indices["GNDVI"])),
            ("08 NDRE", heatmap(indices["NDRE"])),
            ("09 SAVI", heatmap(indices["SAVI"])),
            ("10 OSAVI", heatmap(indices["OSAVI"])),
            ("11 CWSI absolute", heatmap(cwsi, 0, 1)),
            ("12 Thermal C", heatmap(thermal_c, np.nanpercentile(thermal_c, 2), np.nanpercentile(thermal_c, 98))),
            ("13 mask on NDVI", mask_overlay(heatmap(indices["NDVI"]), mask)),
        ],
        STEPS / f"{prefix}_03_indices_cwsi.png",
    )
    return [alignment, segmentation, indices_fig]


def save_segment_outputs(crop, label, capture_id, crop_bands, mask):
    prefix = f"{crop}_{label}_{capture_id}"
    rgb = np.dstack([stretch(crop_bands["red"]), stretch(crop_bands["green"]), stretch(crop_bands["blue"])])

    transparent = np.dstack([rgb, np.where(mask, 255, 0).astype(np.uint8)])
    transparent_path = SEGMENTS / f"{prefix}_plant_segment_transparent.png"
    Image.fromarray(transparent, "RGBA").save(transparent_path)

    white = np.full_like(rgb, 255)
    white[mask] = rgb[mask]
    white_path = SEGMENTS / f"{prefix}_plant_segment_white_bg.png"
    Image.fromarray(white, "RGB").save(white_path)

    binary_path = SEGMENTS / f"{prefix}_plant_mask_binary.png"
    Image.fromarray(np.where(mask, 255, 0).astype(np.uint8), "L").save(binary_path)
    return [transparent_path, white_path, binary_path]


def capture_sets(label_dir):
    files = sorted(label_dir.glob("IMG_*_*.tif"))
    caps = {}
    for file in files:
        m = re.match(r"(IMG_\d+)_([1-7])\.tif$", file.name)
        if m:
            caps.setdefault(m.group(1), {})[int(m.group(2))] = file
    return {cid: group for cid, group in caps.items() if all(b in group for b in BANDS)}


def process_capture(crop, label, capture_id, files):
    green0 = read_tif(files[2])
    full = {"green": green0}
    raw_bands = {"green": green0}
    shifts = {"green": (0, 0)}
    for band_num, band_name in BANDS.items():
        if band_name == "green":
            continue
        arr = read_tif(files[band_num], target_shape=green0.shape)
        raw_bands[band_name] = arr
        shift = estimate_shift(green0, arr, max_shift=80 if band_name != "thermal" else 120)
        full[band_name] = shift_with_nan(arr, *shift)
        shifts[band_name] = shift

    crop_bands = {name: analysis_sample(center_crop(arr)) for name, arr in full.items()}
    mask = green_mask(crop_bands["blue"], crop_bands["green"], crop_bands["red"])
    region_mode, region_labels, comps = choose_regions(mask)

    blue = crop_bands["blue"]
    green = crop_bands["green"]
    red = crop_bands["red"]
    nir1 = crop_bands["nir1"]
    nir2 = crop_bands["nir2_red_edge"]
    thermal_c = thermal_to_celsius(crop_bands["thermal"])

    indices = {
        "NDVI": safe_index(nir1 - red, nir1 + red),
        "GNDVI": safe_index(nir1 - green, nir1 + green),
        "NDRE": safe_index(nir1 - nir2, nir1 + nir2),
        "SAVI": 1.5 * safe_index(nir1 - red, nir1 + red + 0.5),
        "OSAVI": 1.16 * safe_index(nir1 - red, nir1 + red + 0.16),
        "NIRv": nir1 * safe_index(nir1 - red, nir1 + red),
    }
    veg_temp = thermal_c[mask & np.isfinite(thermal_c)]
    if veg_temp.size:
        wet = float(np.nanpercentile(veg_temp, 5))
        dry = float(np.nanpercentile(veg_temp, 95))
    else:
        wet, dry = np.nan, np.nan
    cwsi = np.clip((thermal_c - wet) / (dry - wet), 0, 1) if np.isfinite(wet) and dry > wet else thermal_c * np.nan
    indices["CWSI_absolute"] = cwsi
    indices["Thermal_C"] = thermal_c

    rows = []
    for comp in comps:
        region_mask = mask
        for index_name, arr in indices.items():
            st = stats(arr, region_mask)
            rows.append(
                {
                    "crop": crop,
                    "label": label,
                    "capture_id": capture_id,
                    "region_mode": region_mode,
                    **comp,
                    "index": index_name,
                    "green_mask_fraction_center_crop": float(mask.mean()),
                    "thermal_wet_reference_c": wet,
                    "thermal_dry_reference_c": dry,
                    "shift_blue_yx": str(shifts.get("blue")),
                    "shift_red_yx": str(shifts.get("red")),
                    "shift_nir1_yx": str(shifts.get("nir1")),
                    "shift_nir2_yx": str(shifts.get("nir2_red_edge")),
                    "shift_thermal_yx": str(shifts.get("thermal")),
                    **st,
                }
            )

    preview_path = save_preview(crop, label, capture_id, crop_bands, mask, region_labels, indices["NDVI"], cwsi)
    step_paths = save_step_figures(crop, label, capture_id, raw_bands, full, crop_bands, mask, indices, cwsi, shifts)
    segment_paths = save_segment_outputs(crop, label, capture_id, crop_bands, mask)
    return rows, preview_path, step_paths, segment_paths


def run():
    ensure_dirs()
    all_rows = []
    preview_paths = []
    for crop_dir in sorted([p for p in BASE.iterdir() if p.is_dir()]):
        crop = crop_dir.name
        root = crop_dir / "000"
        if not root.exists():
            continue
        for label_dir in sorted([p for p in root.iterdir() if p.is_dir() and not p.name.startswith("_")]):
            label = label_dir.name.lower()
            if label not in {"control", "test"}:
                continue
            for capture_id, files in sorted(capture_sets(label_dir).items()):
                rows, preview_path, step_paths, segment_paths = process_capture(crop, label, capture_id, files)
                all_rows.extend(rows)
                preview_paths.append(preview_path)
                preview_paths.extend(step_paths)
                preview_paths.extend(segment_paths[:1])

    region_df = pd.DataFrame(all_rows)
    region_csv = OUT / "region_level_indices.csv"
    region_df.to_csv(region_csv, index=False, encoding="utf-8-sig")

    image_df = (
        region_df.groupby(["crop", "label", "capture_id", "index"], as_index=False)
        .agg(
            regions=("region_id", "nunique"),
            mean=("mean", "mean"),
            sd=("mean", "std"),
            median=("median", "median"),
            total_px=("n_px", "sum"),
            green_mask_fraction_center_crop=("green_mask_fraction_center_crop", "first"),
        )
    )
    image_csv = OUT / "image_level_indices.csv"
    image_df.to_csv(image_csv, index=False, encoding="utf-8-sig")

    treatment_df = (
        image_df.groupby(["crop", "label", "index"], as_index=False)
        .agg(captures=("capture_id", "nunique"), mean=("mean", "mean"), sd_across_captures=("mean", "std"), median=("mean", "median"))
    )
    treatment_csv = OUT / "treatment_summary_indices.csv"
    treatment_df.to_csv(treatment_csv, index=False, encoding="utf-8-sig")

    comp = treatment_df[treatment_df["label"].isin(["control", "test"])].pivot_table(
        index=["crop", "index"], columns="label", values="mean", aggfunc="first"
    ).reset_index()
    if "test" in comp.columns and "control" in comp.columns:
        comp["absolute_difference_test_minus_control"] = comp["test"] - comp["control"]
        comp["percent_difference_test_vs_control"] = np.where(comp["control"].abs() > 1e-12, comp["absolute_difference_test_minus_control"] / comp["control"] * 100, np.nan)
    comp_csv = OUT / "test_vs_control_comparison.csv"
    comp.to_csv(comp_csv, index=False, encoding="utf-8-sig")

    fig = draw_comparison(comp)
    workbook = OUT / "labeled_multispectral_analysis.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                ["Crop region", "Only the 0.2-0.8 center range is analyzed: 20% removed from each image edge."],
                ["Alignment", "Bands are resized to Green and translated using FFT phase correlation; this corrects shift but not full perspective distortion."],
                ["Plant mask", "Vegetation mask is Green/ExG based; no NDVI threshold is used for segmentation."],
                ["Regions", "No object/grid split is used; statistics are calculated for the full plant mask area only."],
                ["CWSI", "Absolute CWSI is calculated from thermal C using vegetation-temperature p5/p95 as wet/dry references within each image."],
            ],
            columns=["item", "note"],
        ).to_excel(writer, sheet_name="README", index=False)
        region_df.to_excel(writer, sheet_name="region_level", index=False)
        image_df.to_excel(writer, sheet_name="image_level", index=False)
        treatment_df.to_excel(writer, sheet_name="treatment_summary", index=False)
        comp.to_excel(writer, sheet_name="test_vs_control", index=False)

    wb = load_workbook(workbook)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAD3")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2, 45)
    ws = wb.create_sheet("preview_figures")
    row = 1
    for path in ([fig] if fig else []) + preview_paths[:12]:
        if path and Path(path).exists():
            ws.cell(row=row, column=1, value=Path(path).name)
            img = XLImage(str(path))
            ratio = min(900 / img.width, 1)
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
            ws.add_image(img, f"A{row + 1}")
            row += 22
    wb.save(workbook)

    print(f"Workbook: {workbook}")
    print(f"Region CSV: {region_csv}")
    print(f"Image CSV: {image_csv}")
    print(f"Treatment CSV: {treatment_csv}")
    print(f"Comparison CSV: {comp_csv}")
    print(f"Figures: {FIGURES}")
    print(f"Segments: {SEGMENTS}")
    print(treatment_df.to_string(index=False))


def draw_comparison(comp):
    if comp.empty or "percent_difference_test_vs_control" not in comp.columns:
        return None
    selected = comp[comp["index"].isin(["NDVI", "GNDVI", "NDRE", "CWSI_absolute", "Thermal_C"])].copy()
    selected = selected.sort_values(["crop", "index"])
    rows = selected.to_dict("records")
    width = 1200
    row_h = 32
    top = 60
    height = top + row_h * len(rows) + 35
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    font = ImageFont.load_default()
    draw.text((20, 20), "Test vs Control: vegetation indices and CWSI from labeled MicaSense images", fill=(0, 0, 0), font=font)
    axis = 650
    draw.line([axis, top - 12, axis, height - 25], fill=(90, 90, 90))
    for i, row in enumerate(rows):
        y = top + i * row_h
        pct = row.get("percent_difference_test_vs_control", np.nan)
        draw.text((20, y + 8), f"{row['crop']} | {row['index']}", fill=(0, 0, 0), font=font)
        if np.isfinite(pct):
            scale = 4
            length = int(min(abs(pct) * scale, 460))
            color = (46, 125, 50) if pct >= 0 else (198, 80, 70)
            if pct >= 0:
                draw.rectangle([axis, y + 7, axis + length, y + 24], fill=color)
                tx = axis + length + 8
            else:
                draw.rectangle([axis - length, y + 7, axis, y + 24], fill=color)
                tx = axis - length - 65
            draw.text((tx, y + 8), f"{pct:+.1f}%", fill=(0, 0, 0), font=font)
    out = FIGURES / "test_vs_control_indices_percent_difference.png"
    img.save(out)
    return out


if __name__ == "__main__":
    run()
